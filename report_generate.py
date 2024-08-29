# import tkinter as tk
# from turtle import right
from fpdf import FPDF
from PIL import Image
import pandas as pd
import json

import matplotlib.pyplot as plt
import numpy as np
import openpyxl as op

class SchoolNetwork():
    def __init__(self):
        self.name = ""
        self.schools = []
        self.number_of_students = 0
        self.total_correct_ans = {
            'lp':0, 
            'mat':0,
            'geo':0,
            'his':0,
            'bio':0,
            'ing':0
        }
        self.total_questions = {
            'lp':0, 
            'mat':0,
            'geo':0,
            'his':0,
            'bio':0,
            'ing':0
        }

    def average_correct_ans(self):
        average = {
            'lp':0, 
            'mat':0,
            'geo':0,
            'his':0,
            'bio':0,
            'ing':0
        }
        for key, value in self.total_correct_ans.items():
            average[key] = self.total_correct_ans[key]/self.number_of_students
        return average
    
    def add_school(self, school_name):
        if school_name not in self.schools:
            self.schools.append(school_name)
            
class School():
    def __init__(self):
        self.name = ""
        self.classrooms = []
        self.number_of_students = 0
        self.total_correct_ans = {
            'lp':0, 
            'mat':0,
            'geo':0,
            'his':0,
            'bio':0,
            'ing':0
        }
        self.total_questions = {
            'lp':0, 
            'mat':0,
            'geo':0,
            'his':0,
            'bio':0,
            'ing':0
        }

    def average_correct_ans(self):
        average = {
            'lp':0, 
            'mat':0,
            'geo':0,
            'his':0,
            'bio':0,
            'ing':0
        }
        for key, value in self.total_correct_ans.items():
            average[key] = self.total_correct_ans[key]/self.number_of_students
        return average
    
    def add_classroom(self, clasroom_name):
        if clasroom_name not in self.classrooms:
            self.classrooms.append(clasroom_name)
             
class Classroom():
    def __init__(self):
        self.name = ""
        self.students = []
        self.number_of_students = 0
        self.total_correct_ans = {
            'lp':0, 
            'mat':0,
            'geo':0,
            'his':0,
            'bio':0,
            'ing':0
        }
        self.total_questions = {
            'lp':0, 
            'mat':0,
            'geo':0,
            'his':0,
            'bio':0,
            'ing':0
        }

    def average_correct_ans(self):
        average = {
            'lp':0, 
            'mat':0,
            'geo':0,
            'his':0,
            'bio':0,
            'ing':0
        }
        for key, value in self.total_correct_ans.items():
            average[key] = self.total_correct_ans[key]/self.number_of_students
        return average
            
    def update_total_questions(self, total_questions_to_sum):
        for key, value in self.total_questions:
            self.total_questions[key] += total_questions_to_sum[key]

class Student():
    def __init__(self):
        self.name = ""
        self.group = ""
        self.school_term = ""
        self.grade = ""
        self.school_name = ""
        self.total_correct_ans = {
            "lp":0, 
            "mat":0,
            "geo":0,
            "his":0,
            "bio":0,
            "ing":0
        }
        self.total_questions = {
            "lp":0, 
            "mat":0,
            "geo":0,
            "his":0,
            "bio":0,
            "ing":0
        }
        self.score = 0

workbook = op.load_workbook('modelo_de_dados_copy.xlsx')
ws = workbook.active

subject = ""
competence = ""

score = 0
questions_rigth = 0

school_networks = []
schools = []
classrooms = []
students = []

total_percentage_ans = {
    "lp":0, 
    "mat":0,
    "geo":0,
    "his":0,
    "bio":0,
    "ing":0
}

for irow in range(2, ws.max_row+1):

    student = Student()
    student.name = ws.cell(row=irow, column=1).value
    student.grade = ws.cell(row=irow, column=2).value
    student.group = ws.cell(row=irow, column=3).value
    student.school_term = ws.cell(row=irow, column=4).value
    student.school_name = ws.cell(row=irow, column=5).value

    for icolumn in range(6, ws.max_column+1, 9):
        chose_alternative = ws.cell(row=irow, column=icolumn+7).value
        correct_alternative = ws.cell(row=irow, column=icolumn+8).value
        
        subject = ws.cell(row=irow, column=icolumn).value

        match subject:
            case "LP":
                if (correct_alternative == chose_alternative):
                    student.total_correct_ans["lp"] += 1
                student.total_questions["lp"] += 1
            case "MAT":
                if (correct_alternative == chose_alternative):
                    student.total_correct_ans["mat"] += 1
                student.total_questions["mat"] += 1
            case "GEO":
                if (correct_alternative == chose_alternative):
                    student.total_correct_ans["geo"] += 1
                student.total_questions["geo"] += 1
            case "HIS":
                if (correct_alternative == chose_alternative):
                    student.total_correct_ans["his"] += 1
                student.total_questions["his"] += 1
            case "BIO":
                if (correct_alternative == chose_alternative):
                    student.total_correct_ans["bio"] += 1
                student.total_questions["bio"] += 1
            case "ING":
                if (correct_alternative == chose_alternative) :
                    student.total_correct_ans["ing"] += 1
                student.total_questions["ing"] += 1

# ==============================================================================================

    if not classrooms:
        classroom = Classroom()
        classroom.name = student.group
        classroom.number_of_students += 1
        classroom.students.append(student)
        classroom.total_correct_ans = student.total_correct_ans
        classroom.total_questions = student.total_questions

        classrooms.append(classroom)
    else:  
        found = False

        for classroom in classrooms:
            if classroom.name == student.group:
                classroom.students.append(student)
                classroom.number_of_students += 1

                for key, value in classroom.total_questions.items():
                    classroom.total_questions[key] += student.total_questions[key]

                for key, value in classroom.total_correct_ans.items():
                    classroom.total_correct_ans[key] += student.total_correct_ans[key]
                
                found = True

        if not found:   
            classroom = Classroom()
            classroom.name = student.group
            classroom.number_of_students += 1
            classroom.students.append(student)
            classroom.total_correct_ans = student.total_correct_ans
            classroom.total_questions = student.total_questions

            classrooms.append(classroom)
# ==============================================================================================

    if not schools:
        school = School()
        school.name = student.school_name
        school.number_of_students += 1
        school.add_classroom(student.group)
        school.total_correct_ans = student.total_correct_ans
        school.total_questions = student.total_questions

        schools.append(school)
    else:  
        found = False

        for school in schools:
            if school.name == student.school_name:
                school.number_of_students += 1
                for key, value in school.total_questions.items():
                    school.total_questions[key] += student.total_questions[key]

                for key, value in school.total_correct_ans.items():
                    school.total_correct_ans[key] += student.total_correct_ans[key]
                
                found = True

        if not found:   
            school = School()
            school.name = student.school_name
            school.number_of_students += 1
            school.add_classroom(student.group)
            school.total_correct_ans = student.total_correct_ans
            school.total_questions = student.total_questions

            schools.append(school)
# ==============================================================================================

    if not school_networks:
        school_network = SchoolNetwork()
        school_network.name = "EDUCANDARIO"
        school_network.number_of_students += 1
        school_network.add_school(student.school_name)
        school_network.total_correct_ans = student.total_correct_ans
        school_network.total_questions = student.total_questions

        school_networks.append(school_network)
    else:  
        found = False

        for school_network in school_networks:
            if school_network.name == "EDUCANDARIO":
                school_network.number_of_students += 1
                for key, value in school_network.total_questions.items():
                    school_network.total_questions[key] += student.total_questions[key]

                for key, value in school_network.total_correct_ans.items():
                    school_network.total_correct_ans[key] += student.total_correct_ans[key]
                
                found = True

        if not found:   
            school_network = SchoolNetwork()
            school_network.name = "EDUCANDARIO"
            school_network.number_of_students += 1
            school_network.add_school(student.school_name)
            school_network.total_correct_ans = student.total_correct_ans
            school_network.total_questions = student.total_questions

            school_networks.append(school_network)
        
        # pdf = FPDF(orientation="P", unit="mm", format="A4")
        # pdf.add_page()
        # pdf.set_font("helvetica", "B", 16)
        # pdf.cell(40, 10, "Hello World!")
        # pdf.output(student.name)
        
    # print("Name: ", student.name)
    # print("\n\nTotal questions:")

    for key, value in student.total_questions.items():
        total_percentage_ans[key] = (student.total_correct_ans[key]*100)/(student.total_questions[key]) if student.total_questions[key] else 0
    
    # print("======================\nCorrect ans:")

    for key, value in student.total_correct_ans.items():
        student.score += student.total_correct_ans[key]
    
    # print("======================\nPercentage ans:")

    # for key, value in total_percentage_ans.items():
    #     print(key, " : ", total_percentage_ans[key])
    #     total_percentage_ans[key] = 0

    # print("\nScore: ", student.score)

    students.append(student)

# Choose a student
student_chose = students[25]

# Start create a PDF
pdf = FPDF(orientation="P", unit="mm", format="A4")
pdf.add_page()
pdf.set_font("helvetica", "B", 12)

# Insert follow informations:
# Name
# Grade group / school term         Score
# School
pdf.cell(60, 10, student_chose.name)
pdf.ln()
pdf.cell(60, 10, f"Serie: {student_chose.grade} Turma: {student_chose.group}/ Periodo: {student_chose.school_term}")
pdf.cell(0, 10, f"Nota: {student_chose.score}", align="R")
pdf.ln()
pdf.cell(60, 10, f"Escola: {student_chose.school_name}")
pdf.ln()

################################# Generate graph ##################################################
json_string = json.dumps(students[25].total_correct_ans)
DF = pd.DataFrame(
    data=students[25].total_correct_ans, index=['Acertos', '% Acertos']
    # Convert all data inside dataframe into string type:
).map(str)

COLUMNS = [list(DF)]  # Get list of dataframe columns
ROWS = DF.values.tolist()  # Get list of dataframe rows
DATA = COLUMNS + ROWS  # Combine columns and rows in one list

with pdf.table(
    
    borders_layout="MINIMAL",
    cell_fill_color=200,  # grey
    cell_fill_mode="ROWS",
    line_height=pdf.font_size * 1.5,
    text_align="CENTER",
    width=160,
) as table:
    for data_row in DATA:
        row = table.row()
        for datum in data_row:
            row.cell(datum)

pdf.ln()

################################# Generate graph ##################################################
student_chose = students[25]
classroom_chose = classrooms[2]
school_chose = schools[2]
school_network_chose = school_networks[0]

score_student = [scores for key, scores in student_chose.total_correct_ans.items()]
score_classroom = [scores for key, scores in classroom_chose.average_correct_ans().items()]
score_school = [scores for key, scores in school_chose.average_correct_ans().items()]
score_school_network = [scores for key, scores in school_network_chose.average_correct_ans().items()]

barWidth = 0.15

plt.figure(figsize=(10, 3), dpi=300)

r1 = np.arange(len(score_student))
r2 = [x + barWidth for x in r1]
r3 = [x + barWidth for x in r2]
r4 = [x + barWidth for x in r3]

plt.bar(r1, score_student, color='#00FF55', width=barWidth, label=student_chose.name)
plt.bar(r2, score_classroom, color='#550055', width=barWidth, label='Media da classe')
plt.bar(r3, score_school, color='#FF6655', width=barWidth, label='Media da escola')
plt.bar(r4, score_school_network, color='#FF4455', width=barWidth, label='Media da rede')

plt.xlabel('Gráfico de Provas')
plt.xticks([r + barWidth for r in range(len(score_student))], [key.upper() for key, scores in student_chose.total_correct_ans.items()])
plt.ylabel('Notas')
plt.title('Representacao das notas de 3 alunos em 4 provas')

canvas = plt.gca().figure.canvas
canvas.draw()

img = Image.fromarray(np.asarray(canvas.buffer_rgba()))

pdf.image(img, w=pdf.epw)

pdf.output("student.pdf")

# def exportar_para_pdf():
#     pdf_file = "conteudo_exportado.pdf"

#     # Criar um novo PDF
#     # Converting Figure to an image:
#     canvas = FigureCanvas(fig)
#     canvas.draw()
#     img = Image.fromarray(np.asarray(canvas.buffer_rgba()))

#     pdf = FPDF()
#     pdf.add_page()
#     pdf.image(img, w=pdf.epw)  # Make the image full width
#     pdf.output("matplotlib.pdf")

#     print("PDF exportado com sucesso.")


# # Criar a janela Tkinter
# root = tk.Tk()
# root.title("Exportar para PDF")

# # Criar um botão para acionar a exportação para PDF
# botao_exportar = tk.Button(root, text="Exportar Gráfico para PDF", command=exportar_para_pdf)
# botao_exportar.pack()

# root.mainloop()